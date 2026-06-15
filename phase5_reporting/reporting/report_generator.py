"""
phase5_reporting/report_generator.py
======================================
HFCL EMS — Phase 5 Automated Corporate Reporting Engine
Weekly Network Health Summary  (Excel / openpyxl)

Reads the compiled analytical CSV tables from:
    output/reports/phase4/

Writes a styled executive workbook to:
    output/reports/Weekly_Network_Health_Summary.xlsx

Workbook sheets
---------------
  1. Executive Summary       — dark-navy title, KPI breach overview, operational highlights
  2. Root-Cause Candidates   — Top 10 ranked alarm × KPI correlation pairs
  3. Degradation Windows     — Top Critical Node KPI degradation events (sorted by duration)
  4. Alarm Frequency         — Alarm-type Pareto · NE ranking · category breakdown
  5. KPI Site Ranking        — Per-site breach rate table (all metrics)

Usage
-----
  python -m phase5_reporting.report_generator
  python -m phase5_reporting.report_generator --output-dir output
  python -m phase5_reporting.report_generator --output-dir /data/ems/output
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Optional project logger (falls back to stdlib if project utils unavailable) ─
try:
    from utils.logger import get_logger
    log = get_logger("report_generator")
except ImportError:
    import logging
    log = logging.getLogger("report_generator")
    if not log.handlers:
        logging.basicConfig(level=logging.INFO, format="%(levelname)-8s  %(message)s")


# ═══════════════════════════════════════════════════════════════════════════════
# 0.  DESIGN CONSTANTS — COLOUR PALETTE & TYPOGRAPHY
# ═══════════════════════════════════════════════════════════════════════════════
#
#  openpyxl PatternFill / Font colour values are ARGB hex strings.
#  The first two characters are always "FF" (fully opaque alpha).
#

_C: dict[str, str] = {
    # ── Primary brand palette (aligned with app.py C_* constants) ────────────
    "NAVY_DARK":   "FF1A3D6E",   # deep navy  →  title banners, tab colours
    "NAVY_MID":    "FF2B5EA7",   # mid navy   →  section header bars
    "NAVY_LIGHT":  "FF4472C4",   # pale navy  →  accent accents

    # ── Semantic / alert colours ──────────────────────────────────────────────
    "AMBER":       "FFE59500",   # warning
    "RED":         "FFD64045",   # critical / high-breach
    "GREEN":       "FF3D9970",   # healthy
    "TEAL":        "FF1ABC9C",   # info accent

    # ── Row backgrounds ───────────────────────────────────────────────────────
    "ROW_EVEN":    "FFF2F5FB",   # light blue-grey alternate row
    "ROW_ODD":     "FFFFFFFF",   # white
    "HEADER_ROW":  "FFE8EDF5",   # column-header row (pale blue)
    "CRIT_BG":     "FFFDE8E9",   # red-tint  →  critical rows
    "WARN_BG":     "FFFFF3CD",   # amber-tint →  warning rows
    "GOOD_BG":     "FFE6F4F0",   # green-tint →  healthy rows

    # ── Category-specific pastels (alarm sub-section C) ──────────────────────
    "CAT_RADIO":   "FFE3F0FF",
    "CAT_TRANS":   "FFFDE8E9",
    "CAT_POWER":   "FFFFF3CD",
    "CAT_HW":      "FFE6F4F0",
    "CAT_SW":      "FFF3E5F5",

    # ── Text colours ──────────────────────────────────────────────────────────
    "TEXT_DARK":   "FF1A1A2E",   # body text on light backgrounds
    "TEXT_LIGHT":  "FFFFFFFF",   # text on dark (navy) backgrounds
    "TEXT_MUTED":  "FF6B7280",   # captions / footnotes

    # ── Gridline / border ─────────────────────────────────────────────────────
    "BORDER":      "FFB0BEC5",   # thin grid border colour
    "BORDER_MED":  "FF2B5EA7",   # medium border (section separators)
}

# Sheet tab colours (RGB, no alpha — openpyxl tab_color format)
_TAB: dict[str, str] = {
    "exec":   "1A3D6E",
    "rc":     "D64045",
    "degrad": "E59500",
    "alarm":  "2B5EA7",
    "site":   "3D9970",
}

_FONT_NAME = "Arial"

# Phase 4 CSV file names (relative to reports/phase4/)
_CSV: dict[str, str] = {
    "breach_metric":   "breach_summary_per_metric.csv",
    "breach_site":     "breach_summary_per_site.csv",
    "breach_tech":     "breach_summary_per_technology.csv",
    "breach_temporal": "breach_summary_temporal.csv",
    "degradation":     "kpi_degradation_periods.csv",
    "alarm_type":      "alarm_by_type.csv",
    "alarm_severity":  "alarm_by_severity.csv",
    "alarm_hour":      "alarm_by_hour.csv",
    "alarm_ne":        "alarm_by_ne.csv",
    "alarm_category":  "alarm_by_category.csv",
    "corr_ranked":     "correlation_root_cause_candidates.csv",
    "corr_lag":        "correlation_lag_summary.csv",
    "corr_coo":        "correlation_cooccurrence.csv",
}

# Human-readable KPI display labels (mirrors phase4_analysis and app.py)
_KPI_LABELS: dict[str, str] = {
    "throughput_mbps":           "Throughput (Mbps)",
    "availability_pct":          "Availability (%)",
    "utilization_pct":           "Utilization (%)",
    "latency_ms":                "Latency (ms)",
    "rtwp_dbm":                  "RTWP (dBm)",
    "handover_success_rate_pct": "HO Success Rate (%)",
    "call_drop_rate_pct":        "Call Drop Rate (%)",
    "rach_success_rate_pct":     "RACH Success Rate (%)",
}

# Alarm category → pastel background key
_CAT_FILL: dict[str, str] = {
    "Radio":     _C["CAT_RADIO"],
    "Transport": _C["CAT_TRANS"],
    "Power":     _C["CAT_POWER"],
    "Hardware":  _C["CAT_HW"],
    "Software":  _C["CAT_SW"],
}


# ═══════════════════════════════════════════════════════════════════════════════
# 1.  LOW-LEVEL STYLING HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _fill(hex_argb: str) -> PatternFill:
    """Solid fill from an ARGB hex string."""
    return PatternFill("solid", fgColor=hex_argb)


def _font(
    bold: bool = False,
    size: int = 10,
    color: str = _C["TEXT_DARK"],
    italic: bool = False,
) -> Font:
    return Font(name=_FONT_NAME, bold=bold, size=size, color=color, italic=italic)


def _align(
    h: str = "left",
    v: str = "center",
    wrap: bool = False,
) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _side(style: str = "thin", color: str = _C["BORDER"]) -> Side:
    return Side(style=style, color=color)


def _thin_border() -> Border:
    s = _side("thin", _C["BORDER"])
    return Border(left=s, right=s, top=s, bottom=s)


def _medium_bottom() -> Border:
    """Medium bottom border (section separator)."""
    thin = _side("thin", _C["BORDER"])
    med  = _side("medium", _C["BORDER_MED"])
    return Border(left=thin, right=thin, top=thin, bottom=med)


# ── Cell-level write helpers ──────────────────────────────────────────────────

def _write(
    ws,
    row: int,
    col: int,
    value,
    *,
    fill:   Optional[PatternFill] = None,
    font:   Optional[Font]        = None,
    align:  Optional[Alignment]   = None,
    border: Optional[Border]      = None,
    fmt:    str                   = "General",
):
    """Write a single cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill            = fill
    if font:   cell.font            = font
    if align:  cell.alignment       = align
    if border: cell.border          = border
    if fmt != "General":
        cell.number_format = fmt
    return cell


def _merge_write(
    ws,
    row: int,
    col_start: int,
    col_end: int,
    value: str,
    *,
    fill:   Optional[PatternFill] = None,
    font:   Optional[Font]        = None,
    align:  Optional[Alignment]   = None,
    height: Optional[float]       = None,
):
    """Merge a row span and write a value into it."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end,
    )
    cell = ws.cell(row=row, column=col_start, value=value)
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    if height is not None:
        ws.row_dimensions[row].height = height
    return cell


# ── Composite helpers ─────────────────────────────────────────────────────────

def _title_banner(
    ws,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int,
    text: str,
    font_size: int = 18,
):
    """Write a multi-row dark-navy title banner (merged cells)."""
    for r in range(row_start, row_end + 1):
        ws.row_dimensions[r].height = 20
    ws.merge_cells(
        start_row=row_start, start_column=col_start,
        end_row=row_end,     end_column=col_end,
    )
    cell = ws.cell(row=row_start, column=col_start, value=text)
    cell.fill      = _fill(_C["NAVY_DARK"])
    cell.font      = Font(name=_FONT_NAME, bold=True, size=font_size,
                          color=_C["TEXT_LIGHT"])
    cell.alignment = _align("center", "center")
    return cell


def _section_bar(ws, row: int, col_start: int, col_end: int, label: str,
                 height: float = 22):
    """Render a medium-navy section-header bar (merged, left-aligned label)."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end,
    )
    cell = ws.cell(row=row, column=col_start, value=label)
    cell.fill      = _fill(_C["NAVY_MID"])
    cell.font      = _font(bold=True, size=11, color=_C["TEXT_LIGHT"])
    cell.alignment = _align("left", "center")
    ws.row_dimensions[row].height = height
    return cell


def _col_headers(ws, row: int, col_start: int, labels: list[str],
                 height: float = 28):
    """Write a row of column-header cells on the HEADER_ROW background."""
    for i, label in enumerate(labels):
        _write(
            ws, row, col_start + i, label,
            fill=_fill(_C["HEADER_ROW"]),
            font=_font(bold=True, size=9, color=_C["NAVY_DARK"]),
            align=_align("center", "center", wrap=True),
            border=_thin_border(),
        )
    ws.row_dimensions[row].height = height


def _set_col_widths(ws, mapping: dict[int, float]):
    """Set column widths by column index (1-based)."""
    for col_idx, width in mapping.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─ Safe scalar extraction from a DataFrame ────────────────────────────────────

def _scalar(df: Optional[pd.DataFrame], col: str, default="N/A"):
    """Return df.iloc[0][col] safely, or default on any failure."""
    if df is None or df.empty or col not in df.columns:
        return default
    v = df.iloc[0][col]
    return v if pd.notna(v) else default


def _kpi_label(raw: str) -> str:
    """Map a raw metric column name to its display label."""
    return _KPI_LABELS.get(raw, raw)


# ═══════════════════════════════════════════════════════════════════════════════
# 2.  CSV DATA LOADER
# ═══════════════════════════════════════════════════════════════════════════════

def _load_csvs(reports_dir: Path) -> dict[str, Optional[pd.DataFrame]]:
    """
    Load every Phase 4 CSV report into a keyed dict.

    Missing files are logged as warnings and stored as ``None``.
    Each sheet-builder handles ``None`` gracefully with a placeholder message.
    """
    data: dict[str, Optional[pd.DataFrame]] = {}
    for key, filename in _CSV.items():
        path = reports_dir / filename
        if path.exists():
            try:
                data[key] = pd.read_csv(path, low_memory=False)
                log.info(f"  ✓  {filename:<50}  ({len(data[key]):>6,} rows)")
            except Exception as exc:
                log.warning(f"  ✗  {filename}  →  parse error: {exc}")
                data[key] = None
        else:
            log.warning(f"  –  {filename:<50}  (not found — will render blank)")
            data[key] = None
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# 3.  SHEET 1 — EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def _build_executive_summary(ws, data: dict, generated_at: str):
    """
    Construct the Executive Summary sheet.

    Layout (row numbers are approximate — content-length can shift them):
    ───────────────────────────────────────────────────────────────────
    Rows  1–3  │  Dark-navy title banner
    Row   4    │  Spacer
    Row   5    │  Metadata strip  (generated timestamp, data source)
    Row   6    │  Spacer
    Row   7    │  Section bar — "SECTION 1 – KPI NETWORK HEALTH"
    Row   8    │  Column headers for KPI breach table
    Rows  9-N  │  KPI breach per-metric rows (colour-coded by breach rate)
    Row   N+1  │  Spacer
    Row   N+2  │  Section bar — "SECTION 2 – ALARM VOLUME OVERVIEW"
    Row   N+3  │  Column headers for alarm severity table
    Rows  N+4  │  Alarm-by-severity rows
    Row   M+1  │  Spacer
    Row   M+2  │  Section bar — "SECTION 3 – KEY OPERATIONAL HIGHLIGHTS"
    Row   M+3  │  Column headers for highlights
    Rows  M+4  │  Highlight KV rows
    Row   last │  Footnote disclaimer
    ───────────────────────────────────────────────────────────────────
    """
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = _TAB["exec"]
    ws.sheet_view.showGridLines  = True

    # Column widths: A(margin) B..J(content) K(margin)
    _set_col_widths(ws, {
        1: 3,   2: 26,  3: 22,  4: 14,  5: 14,
        6: 14,  7: 18,  8: 14,  9: 14, 10: 14,  11: 3,
    })

    D  = 2   # data starts at column B (index 2)
    NC = 9   # number of content columns (B..J)

    # ── Rows 1–3: Title banner ────────────────────────────────────────────────
    _title_banner(ws, 1, 3, 1, D + NC,
                  "HFCL EMS — Weekly Network Health Summary",
                  font_size=18)

    # ── Row 4: spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Row 5: metadata strip ─────────────────────────────────────────────────
    ws.row_dimensions[5].height = 16
    _write(ws, 5, D,   "Generated:",
           font=_font(bold=True, size=9, color=_C["TEXT_MUTED"]),
           align=_align("right", "center"))
    _write(ws, 5, D+1, generated_at,
           font=_font(size=9, color=_C["TEXT_MUTED"]),
           align=_align("left", "center"))
    _write(ws, 5, D+4, "Data source:",
           font=_font(bold=True, size=9, color=_C["TEXT_MUTED"]),
           align=_align("right", "center"))
    ws.merge_cells(f"{get_column_letter(D+5)}5:{get_column_letter(D+NC-1)}5")
    _write(ws, 5, D+5, "output/reports/phase4/  (Phase 4 Analysis Engine)",
           font=_font(size=9, italic=True, color=_C["TEXT_MUTED"]),
           align=_align("left", "center"))

    # ── Row 6: spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 10

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 1 — KPI Breach Summary
    # ─────────────────────────────────────────────────────────────────────────
    row = 7
    _section_bar(ws, row, D, D + NC - 1,
                 "  SECTION 1 — KPI NETWORK HEALTH AT A GLANCE")
    row += 1

    kpi_hdrs = [
        "KPI Metric", "Breach Direction", "SLA Threshold",
        "Breach Count", "Breach Rate (%)", "Total Observations",
    ]
    _col_headers(ws, row, D, kpi_hdrs, height=26)
    row += 1

    breach_df = data.get("breach_metric")
    if breach_df is not None and not breach_df.empty:
        for i, (_, r) in enumerate(breach_df.iterrows()):
            metric    = r.get("metric", "")
            disp      = _kpi_label(r.get("display_name", metric) if "display_name" in r.index else metric)
            direction = r.get("breach_direction", "")
            threshold = r.get("threshold", "")
            b_count   = r.get("breach_count", "")
            b_rate    = r.get("breach_rate_pct", "")
            total_obs = r.get("total_observations", "")

            # Row background: red ≥20 %, amber 10–20 %, otherwise alternating
            if isinstance(b_rate, (int, float)) and not pd.isna(b_rate):
                bg = (_C["CRIT_BG"] if b_rate >= 20 else
                      _C["WARN_BG"] if b_rate >= 10 else
                      (_C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]))
            else:
                bg = _C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]

            # Column B: metric name (bold navy)
            _write(ws, row, D, disp,
                   fill=_fill(bg),
                   font=_font(bold=True, size=9, color=_C["NAVY_DARK"]),
                   align=_align("left", "center"),
                   border=_thin_border())

            # Columns C..G: data values
            vals = [direction, threshold, b_count, b_rate, total_obs]
            fmts = ["General", "#,##0.00", "#,##0", "0.00", "#,##0"]
            for j, (v, fmt) in enumerate(zip(vals, fmts)):
                _write(ws, row, D+1+j, v,
                       fill=_fill(bg),
                       font=_font(size=9),
                       align=_align("center", "center"),
                       border=_thin_border(),
                       fmt=fmt)

            ws.row_dimensions[row].height = 17
            row += 1
    else:
        ws.merge_cells(f"{get_column_letter(D)}{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D,
               "[ KPI breach data not available — run Phase 4 first ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]),
               align=_align("left"))
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1   # spacer

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 2 — Alarm Volume Overview
    # ─────────────────────────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  SECTION 2 — ALARM VOLUME OVERVIEW  (by Severity Level)")
    row += 1

    sev_hdrs = [
        "Severity", "Severity Rank", "Count", "% of Total",
        "Avg Duration (min)", "Median Duration (min)", "Affected Sites",
    ]
    _col_headers(ws, row, D, sev_hdrs, height=24)
    row += 1

    _SEV_ROW_BG = {
        "Critical": _C["CRIT_BG"],
        "Major":    _C["WARN_BG"],
        "Minor":    _C["ROW_EVEN"],
        "Warning":  _C["GOOD_BG"],
    }
    _SEV_TXT = {
        "Critical": _C["RED"],
        "Major":    "FF8B4000",   # dark amber-brown
        "Minor":    _C["NAVY_DARK"],
        "Warning":  "FF1B5E20",   # dark green
    }

    sev_df = data.get("alarm_severity")
    if sev_df is not None and not sev_df.empty:
        for _, r in sev_df.iterrows():
            sev = r.get("severity", "")
            bg  = _SEV_ROW_BG.get(sev, _C["ROW_ODD"])
            tc  = _SEV_TXT.get(sev, _C["TEXT_DARK"])

            _write(ws, row, D, sev,
                   fill=_fill(bg),
                   font=_font(bold=True, size=9, color=tc),
                   align=_align("center", "center"),
                   border=_thin_border())

            vals = [
                r.get("severity_rank", ""),
                r.get("count", ""),
                r.get("pct_of_total", ""),
                r.get("avg_duration_min", ""),
                r.get("median_duration_min", ""),
                r.get("affected_sites", ""),
            ]
            fmts = ["General", "#,##0", "0.00", "0.0", "0.0", "#,##0"]
            for j, (v, fmt) in enumerate(zip(vals, fmts)):
                _write(ws, row, D+1+j, v,
                       fill=_fill(bg),
                       font=_font(size=9),
                       align=_align("center", "center"),
                       border=_thin_border(),
                       fmt=fmt)

            ws.row_dimensions[row].height = 17
            row += 1
    else:
        ws.merge_cells(f"{get_column_letter(D)}{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D,
               "[ Alarm severity data not available — run Phase 4 first ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]),
               align=_align("left"))
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1   # spacer

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 3 — Key Operational Highlights
    # ─────────────────────────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  SECTION 3 — KEY OPERATIONAL HIGHLIGHTS")
    row += 1

    # Build highlight items dynamically from available data
    highlights: list[tuple[str, str, str]] = []

    bm = data.get("breach_metric")
    if bm is not None and not bm.empty:
        worst_m    = _kpi_label(_scalar(bm, "metric"))
        worst_rate = _scalar(bm, "breach_rate_pct")
        detail     = (f"{worst_rate:.1f}% breach rate"
                      if isinstance(worst_rate, (int, float)) else str(worst_rate))
        highlights.append(("Worst KPI Metric",              worst_m,      detail))

    bs = data.get("breach_site")
    if bs is not None and not bs.empty and "any_breach_pct" in bs.columns:
        top_site  = bs.sort_values("any_breach_pct", ascending=False).iloc[0]
        ne_id     = top_site.get("network_element_id", "N/A")
        site_rate = top_site.get("any_breach_pct", "N/A")
        detail    = (f"{site_rate:.1f}% any-KPI breach rate"
                     if isinstance(site_rate, (int, float)) else str(site_rate))
        highlights.append(("Highest-Breach Site",            ne_id,        detail))

    an = data.get("alarm_ne")
    if an is not None and not an.empty:
        noisiest = _scalar(an, "network_element_id")
        noise_ct = _scalar(an, "total_alarms")
        detail   = (f"{int(noise_ct):,} total alarms"
                    if isinstance(noise_ct, (int, float)) else str(noise_ct))
        highlights.append(("Noisiest Network Element",       noisiest,     detail))

    at_df = data.get("alarm_type")
    if at_df is not None and not at_df.empty:
        top_type = _scalar(at_df, "alarm_type")
        top_pct  = _scalar(at_df, "pct_of_total")
        detail   = (f"{top_pct:.1f}% of total alarm volume"
                    if isinstance(top_pct, (int, float)) else str(top_pct))
        highlights.append(("Top Alarm Type",                 top_type,     detail))

    rc = data.get("corr_ranked")
    if rc is not None and not rc.empty:
        top_alarm = _scalar(rc, "alarm_type")
        top_kpi   = _kpi_label(_scalar(rc, "kpi_metric"))
        top_score = _scalar(rc, "combined_score")
        detail    = (f"Combined correlation score: {top_score:.3f}"
                     if isinstance(top_score, (int, float)) else "")
        highlights.append(("Top Root-Cause Pair",
                            f"{top_alarm}  →  {top_kpi}",  detail))

    deg = data.get("degradation")
    if deg is not None and not deg.empty:
        if "duration_hours" in deg.columns:
            top_deg = deg.sort_values("duration_hours", ascending=False).iloc[0]
        else:
            top_deg = deg.iloc[0]
        worst_dur = top_deg.get("duration_hours", "N/A")
        deg_ne    = top_deg.get("network_element_id", "N/A")
        deg_met   = _kpi_label(top_deg.get("metric", ""))
        detail    = (f"{worst_dur:.1f}h sustained breach window"
                     if isinstance(worst_dur, (int, float)) else "")
        highlights.append(("Longest Degradation Window",
                            f"{deg_ne} — {deg_met}",        detail))

    hl_hdrs = ["Highlight", "Value", "Detail"]
    _col_headers(ws, row, D, hl_hdrs, height=22)
    row += 1

    for idx, (label, value, detail) in enumerate(highlights):
        bg = _C["ROW_EVEN"] if idx % 2 == 0 else _C["ROW_ODD"]
        for j, (v, is_bold) in enumerate([(label, True), (value, False), (detail, False)]):
            _write(ws, row, D+j, str(v),
                   fill=_fill(bg),
                   font=_font(bold=is_bold, size=9,
                              color=_C["NAVY_DARK"] if is_bold else _C["TEXT_DARK"]),
                   align=_align("left", "center"),
                   border=_thin_border())
        ws.row_dimensions[row].height = 16
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1   # spacer

    # ── Footnote ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"{get_column_letter(D)}{row}:{get_column_letter(D+NC-1)}{row}")
    fn = ws.cell(row=row, column=D,
                 value=(
                     "Auto-generated by HFCL EMS Phase 5 Reporting Engine.  "
                     "Data sourced from Phase 4 CSV exports.  "
                     "Row colours: Red ≥20% breach rate · Amber 10–20% · "
                     "Alternating blue-grey otherwise.  "
                     "Interactive drill-down: run  streamlit run app.py"
                 ))
    fn.font      = _font(size=8, italic=True, color=_C["TEXT_MUTED"])
    fn.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[row].height = 30

    ws.freeze_panes = "B9"


# ═══════════════════════════════════════════════════════════════════════════════
# 4.  SHEET 2 — TOP 10 ROOT-CAUSE ALARM CANDIDATES
# ═══════════════════════════════════════════════════════════════════════════════

def _build_root_cause_sheet(ws, data: dict):
    """
    Populate the Root-Cause Candidates worksheet.

    Source: correlation_root_cause_candidates.csv  (top 10 rows)

    Columns
    -------
    # | Alarm Type | KPI Metric | Combined Score | Lag Score |
    Co-occurrence Rate | Peak Lag (h) | Degrades KPI |
    Alarm Events | Co-occ Count | Interpretation
    """
    ws.title = "Root-Cause Candidates"
    ws.sheet_properties.tabColor = _TAB["rc"]
    ws.sheet_view.showGridLines  = True

    _set_col_widths(ws, {
        1:  3,   # margin
        2:  6,   # rank
        3:  30,  # alarm type
        4:  26,  # KPI metric
        5:  16,  # combined score
        6:  14,  # lag score
        7:  18,  # co-occ rate
        8:  14,  # peak lag
        9:  14,  # degrades KPI
        10: 14,  # alarm events
        11: 14,  # co-occ count
        12: 46,  # interpretation
        13: 3,   # margin
    })

    D  = 2   # data start column (B)
    NC = 11  # B..L

    # ── Rows 1–2: title banner ────────────────────────────────────────────────
    _title_banner(ws, 1, 2, 1, D + NC + 1,
                  "HFCL EMS — Top 10 Detected Root-Cause Alarm Candidates",
                  font_size=15)

    # ── Row 3: sub-title caption ──────────────────────────────────────────────
    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"B3:{get_column_letter(D+NC-1)}3")
    cap = ws["B3"]
    cap.value = (
        "Ranked by combined lag-analysis × co-occurrence score (0–1). "
        "High combined_score → strong statistical evidence of root-cause relationship. "
        "Row colour: Red ≥0.70 (High priority)  ·  Amber 0.40–0.69 (Medium)  ·  Alternating otherwise (Low)."
    )
    cap.font      = _font(size=9, italic=True, color=_C["TEXT_MUTED"])
    cap.alignment = _align("left", "center", wrap=True)

    ws.row_dimensions[4].height = 8   # spacer

    # ── Row 5: column headers ─────────────────────────────────────────────────
    hdrs = [
        "#", "Alarm Type", "KPI Metric",
        "Combined\nScore", "Lag\nScore", "Co-occ.\nRate",
        "Peak Lag\n(hours)", "Degrades\nKPI",
        "Alarm\nEvents", "Co-occ.\nCount", "Interpretation",
    ]
    _col_headers(ws, 5, D, hdrs, height=34)

    # ── Rows 6+: data ─────────────────────────────────────────────────────────
    rc_df = data.get("corr_ranked")
    row   = 6

    if rc_df is not None and not rc_df.empty:
        top10 = rc_df.head(10).reset_index(drop=True)

        for i, (_, r) in enumerate(top10.iterrows()):
            comb_s = r.get("combined_score", None)

            # Row background by score tier
            if isinstance(comb_s, float) and not pd.isna(comb_s):
                bg = (_C["CRIT_BG"] if comb_s >= 0.70 else
                      _C["WARN_BG"] if comb_s >= 0.40 else
                      (_C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]))
            else:
                bg = _C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]

            rank     = r.get("rank", i + 1)
            alarm_t  = r.get("alarm_type", "")
            kpi_raw  = r.get("kpi_metric", "")
            kpi_disp = _kpi_label(r.get("kpi_display", kpi_raw) if "kpi_display" in r.index else kpi_raw)
            lag_s    = r.get("lag_score", None)
            coo_r    = r.get("cooccurrence_rate", None)
            peak_lag = r.get("peak_lag_hours", None)
            degrades = r.get("degrades_kpi", "")
            n_ev     = r.get("n_alarm_events", "")
            coo_cnt  = r.get("cooccurrence_count", "")
            interp   = r.get("interpretation", "")

            vals = [rank, alarm_t, kpi_disp, comb_s, lag_s, coo_r,
                    peak_lag, str(degrades), n_ev, coo_cnt, interp]
            fmts = ["General", "General", "General",
                    "0.000", "0.000", "0.000",
                    "0.0", "General", "#,##0", "#,##0", "General"]

            for j, (v, fmt) in enumerate(zip(vals, fmts)):
                bold   = j in (1, 2)
                txt_c  = _C["NAVY_DARK"] if j in (1, 2) else _C["TEXT_DARK"]
                halign = "left" if j in (1, 2, 10) else "center"
                wrap   = (j == 10)

                cell = ws.cell(row=row, column=D+j, value=v)
                cell.fill      = _fill(bg)
                cell.font      = _font(bold=bold, size=9, color=txt_c)
                cell.alignment = _align(halign, "center", wrap=wrap)
                cell.border    = _thin_border()
                if fmt != "General":
                    cell.number_format = fmt

            # Taller rows for long interpretation text
            ws.row_dimensions[row].height = (
                32 if interp and len(str(interp)) > 70 else 17
            )
            row += 1

    else:
        ws.merge_cells(f"B{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D,
               "[ correlation_root_cause_candidates.csv not found — run Phase 4 correlation analysis ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]),
               align=_align("left"))
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1  # spacer

    # ── Score legend ──────────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  SCORE INTERPRETATION LEGEND")
    row += 1

    legend = [
        (_C["CRIT_BG"], "Combined Score ≥ 0.70 — HIGH priority",
         "Strong lag + co-occurrence evidence. Immediate investigation recommended. "
         "Alarm type highly likely to be a root cause of the linked KPI degradation."),
        (_C["WARN_BG"], "Combined Score 0.40 – 0.69 — MEDIUM priority",
         "Moderate correlation signal. Schedule targeted review within the maintenance window."),
        (_C["ROW_EVEN"], "Combined Score < 0.40 — LOW priority",
         "Weak or inconclusive signal. Monitor passively; revisit if other evidence emerges."),
    ]
    for bg_c, label, desc in legend:
        _write(ws, row, D, label,
               fill=_fill(bg_c),
               font=_font(bold=True, size=9),
               align=_align("left", "center"),
               border=_thin_border())
        ws.merge_cells(f"{get_column_letter(D+1)}{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D+1, desc,
               fill=_fill(bg_c),
               font=_font(size=9),
               align=_align("left", "center", wrap=True),
               border=_thin_border())
        ws.row_dimensions[row].height = 28
        row += 1

    ws.freeze_panes = "B6"


# ═══════════════════════════════════════════════════════════════════════════════
# 5.  SHEET 3 — TOP CRITICAL NODE DEGRADATION WINDOWS
# ═══════════════════════════════════════════════════════════════════════════════

def _build_degradation_sheet(ws, data: dict):
    """
    Populate the Critical Node KPI Degradation Windows worksheet.

    Source: kpi_degradation_periods.csv  (sorted by duration_hours DESC, top 40)

    Columns
    -------
    # | Network Element | KPI Metric | Window Start | Window End |
    Duration (h) | Duration Periods | Worst Value | Mean Value |
    Threshold | Breach Direction | Severity Class
    """
    ws.title = "Degradation Windows"
    ws.sheet_properties.tabColor = _TAB["degrad"]
    ws.sheet_view.showGridLines  = True

    _set_col_widths(ws, {
        1:  3,   # margin
        2:  6,   # #
        3:  22,  # network element
        4:  24,  # KPI metric
        5:  20,  # start
        6:  20,  # end
        7:  14,  # duration (h)
        8:  12,  # duration periods
        9:  14,  # worst value
        10: 14,  # mean value
        11: 12,  # threshold
        12: 14,  # breach direction
        13: 16,  # severity class
        14: 3,   # margin
    })

    D  = 2   # data start column
    NC = 12  # B..M

    # ── Rows 1–2: title banner ────────────────────────────────────────────────
    _title_banner(ws, 1, 2, 1, D + NC + 1,
                  "HFCL EMS — Top Critical Node KPI Degradation Windows",
                  font_size=15)

    # ── Row 3: sub-title ──────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"B3:{get_column_letter(D+NC-1)}3")
    cap = ws["B3"]
    cap.value = (
        "Sustained KPI degradation windows: ≥ 4 consecutive 15-min breach intervals (≥ 1 hour).  "
        "Sorted by duration descending.  "
        "Severity class: Critical ≥ 6h · Major 2–6h · Minor < 2h."
    )
    cap.font      = _font(size=9, italic=True, color=_C["TEXT_MUTED"])
    cap.alignment = _align("left", "center", wrap=True)

    ws.row_dimensions[4].height = 8   # spacer

    # ── Load and sort data ────────────────────────────────────────────────────
    deg_df = data.get("degradation")
    TOP_N  = 40

    if deg_df is not None and not deg_df.empty:
        if "duration_hours" in deg_df.columns:
            deg_df = deg_df.sort_values("duration_hours", ascending=False)
        display_df = deg_df.head(TOP_N).reset_index(drop=True)
    else:
        display_df = pd.DataFrame()

    # ── Row 5: column headers ─────────────────────────────────────────────────
    hdrs = [
        "#", "Network\nElement", "KPI Metric",
        "Window Start", "Window End",
        "Duration\n(hours)", "Duration\n(Periods)",
        "Worst\nValue", "Mean\nValue",
        "Threshold", "Breach\nDirection", "Severity\nClass",
    ]
    _col_headers(ws, 5, D, hdrs, height=34)

    row = 6

    if not display_df.empty:
        for i, (_, r) in enumerate(display_df.iterrows()):
            dur_h = r.get("duration_hours", None)

            # Classify and colour by duration
            if isinstance(dur_h, (int, float)) and not pd.isna(dur_h):
                if dur_h >= 6:
                    sev_class = "Critical"
                    bg        = _C["CRIT_BG"]
                elif dur_h >= 2:
                    sev_class = "Major"
                    bg        = _C["WARN_BG"]
                else:
                    sev_class = "Minor"
                    bg        = _C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]
            else:
                sev_class = "Unknown"
                bg        = _C["ROW_ODD"]

            ne     = r.get("network_element_id", "")
            metric = _kpi_label(r.get("metric", ""))
            start  = str(r.get("start", ""))
            end    = str(r.get("end", ""))
            dur_p  = r.get("duration_periods", "")
            worst  = r.get("worst_value", "")
            mean   = r.get("mean_value", "")
            thresh = r.get("threshold", "")
            direc  = r.get("breach_direction", "")

            sev_txt = {
                "Critical": _C["RED"],
                "Major":    "FF8B4000",
                "Minor":    _C["NAVY_DARK"],
            }.get(sev_class, _C["TEXT_DARK"])

            vals = [i+1, ne, metric, start, end,
                    dur_h, dur_p, worst, mean, thresh, direc, sev_class]
            fmts = ["General", "General", "General", "General", "General",
                    "0.00", "#,##0", "0.0000", "0.0000", "0.00",
                    "General", "General"]

            for j, (v, fmt) in enumerate(zip(vals, fmts)):
                col_idx = D + j
                is_sev  = (j == 11)
                bold    = j in (1, 2) or is_sev
                txt_c   = sev_txt if is_sev else (_C["NAVY_DARK"] if j in (1, 2) else _C["TEXT_DARK"])
                halign  = "left" if j in (1, 2) else "center"

                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.fill      = _fill(bg)
                cell.font      = _font(bold=bold, size=9, color=txt_c)
                cell.alignment = _align(halign, "center")
                cell.border    = _thin_border()
                if fmt != "General":
                    cell.number_format = fmt

            ws.row_dimensions[row].height = 16
            row += 1

    else:
        ws.merge_cells(f"B{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D,
               "[ kpi_degradation_periods.csv not found — run Phase 4 KPI trending analysis ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]),
               align=_align("left"))
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1  # spacer

    # ── Severity legend ───────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  DEGRADATION SEVERITY CLASSIFICATION")
    row += 1

    legend = [
        (_C["CRIT_BG"], "Critical  —  Duration ≥ 6 hours",
         "Immediate escalation required. High risk of customer-impacting SLA breach."),
        (_C["WARN_BG"], "Major  —  Duration 2–6 hours",
         "Prompt investigation within current maintenance cycle."),
        (_C["ROW_EVEN"], "Minor  —  Duration < 2 hours",
         "Schedule review. Monitor for recurrence before next reporting cycle."),
    ]
    for bg_c, label, desc in legend:
        _write(ws, row, D, label,
               fill=_fill(bg_c),
               font=_font(bold=True, size=9),
               align=_align("left", "center"),
               border=_thin_border())
        ws.merge_cells(f"{get_column_letter(D+1)}{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D+1, desc,
               fill=_fill(bg_c),
               font=_font(size=9),
               align=_align("left", "center"),
               border=_thin_border())
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "B6"


# ═══════════════════════════════════════════════════════════════════════════════
# 6.  SHEET 4 — ALARM FREQUENCY SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def _build_alarm_frequency_sheet(ws, data: dict):
    """
    Populate the Alarm Frequency Summary worksheet.

    Sub-section A  —  Alarm-by-Type Pareto (alarm_by_type.csv)
    Sub-section B  —  Network Element Alarm Ranking (alarm_by_ne.csv)
    Sub-section C  —  Alarm Category Breakdown (alarm_by_category.csv)
    """
    ws.title = "Alarm Frequency"
    ws.sheet_properties.tabColor = _TAB["alarm"]
    ws.sheet_view.showGridLines  = True

    _set_col_widths(ws, {
        1:  3,   # margin
        2:  6,   # rank
        3:  30,  # name col
        4:  18,  # category / region
        5:  12,  # count / tech
        6:  12,  # % total / total
        7:  14,  # cumulative % / critical
        8:  14,  # avg duration / major
        9:  14,  # severity mix / minor
        10: 12,  # warning
        11: 14,  # avg duration (NE)
        12: 3,   # margin
    })

    D  = 2    # data start column
    NC = 10   # B..K

    # ── Title banner ──────────────────────────────────────────────────────────
    _title_banner(ws, 1, 2, 1, D + NC + 1,
                  "HFCL EMS — Alarm Frequency Analysis Summary",
                  font_size=15)
    ws.row_dimensions[3].height = 8   # spacer
    row = 4

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-SECTION A: Alarm-by-Type Pareto
    # ─────────────────────────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  A.  ALARM-TYPE PARETO TABLE  (Top 20 by Volume · Sorted Descending)")
    row += 1

    _col_headers(ws, row, D,
                 ["#", "Alarm Type", "Category", "Count",
                  "% of Total", "Cumulative %", "Avg Duration (min)", "Severity Mix"],
                 height=26)
    row += 1

    at_df = data.get("alarm_type")
    if at_df is not None and not at_df.empty:
        for i, (_, r) in enumerate(at_df.head(20).iterrows()):
            cum_pct = r.get("cumulative_pct", None)

            # Rows within the 80% Pareto boundary get white; beyond get even-grey
            if isinstance(cum_pct, (int, float)) and not pd.isna(cum_pct):
                bg = _C["ROW_ODD"] if cum_pct <= 80 else _C["ROW_EVEN"]
            else:
                bg = _C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]

            vals = [
                i+1,
                r.get("alarm_type", ""),
                r.get("alarm_category", ""),
                r.get("count", ""),
                r.get("pct_of_total", ""),
                cum_pct,
                r.get("avg_duration_min", ""),
                str(r.get("severity_mix", "")),
            ]
            fmts = ["General", "General", "General", "#,##0", "0.00", "0.00", "0.0", "General"]
            aligns = ["center", "left", "left", "center", "center", "center", "center", "left"]

            for j, (v, fmt, ha) in enumerate(zip(vals, fmts, aligns)):
                bold  = (j == 1)
                txt_c = _C["NAVY_DARK"] if j == 1 else _C["TEXT_DARK"]
                cell  = ws.cell(row=row, column=D+j, value=v)
                cell.fill      = _fill(bg)
                cell.font      = _font(bold=bold, size=9, color=txt_c)
                cell.alignment = _align(ha, "center", wrap=(j == 7))
                cell.border    = _thin_border()
                if fmt != "General":
                    cell.number_format = fmt

            ws.row_dimensions[row].height = 16
            row += 1

        # Pareto boundary annotation
        ws.merge_cells(f"B{row}:{get_column_letter(D+NC-1)}{row}")
        ann = ws.cell(row=row, column=D,
                      value="▲  Rows above the grey-shaded boundary contribute ≤ 80% of total alarm volume "
                            "(Pareto 80/20 principle).  Prioritise these alarm types for root-cause elimination.")
        ann.font      = _font(size=8, italic=True, color=_C["TEXT_MUTED"])
        ann.alignment = _align("left", "center")
        ws.row_dimensions[row].height = 13
        row += 1
    else:
        ws.merge_cells(f"B{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D, "[ alarm_by_type.csv not found ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]), align=_align("left"))
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1   # spacer

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-SECTION B: Network Element Alarm Ranking
    # ─────────────────────────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  B.  NETWORK ELEMENT ALARM RANKING  (Top 20 · Sorted by Total Alarms)")
    row += 1

    _col_headers(ws, row, D,
                 ["#", "Network Element", "Region", "Technology",
                  "Total Alarms", "Critical", "Major", "Minor", "Warning",
                  "Avg Duration (min)"],
                 height=26)
    row += 1

    ne_df = data.get("alarm_ne")
    if ne_df is not None and not ne_df.empty:
        for i, (_, r) in enumerate(ne_df.head(20).iterrows()):
            bg   = _C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]
            vals = [
                i+1,
                r.get("network_element_id", ""),
                r.get("region", ""),
                r.get("technology", ""),
                r.get("total_alarms", ""),
                r.get("critical_count", ""),
                r.get("major_count", ""),
                r.get("minor_count", ""),
                r.get("warning_count", ""),
                r.get("avg_duration_min", ""),
            ]
            fmts   = ["General", "General", "General", "General",
                      "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "0.0"]
            aligns = ["center", "left", "left", "center",
                      "center", "center", "center", "center", "center", "center"]

            for j, (v, fmt, ha) in enumerate(zip(vals, fmts, aligns)):
                bold  = (j == 1)
                txt_c = _C["NAVY_DARK"] if j == 1 else _C["TEXT_DARK"]
                cell  = ws.cell(row=row, column=D+j, value=v)
                cell.fill      = _fill(bg)
                cell.font      = _font(bold=bold, size=9, color=txt_c)
                cell.alignment = _align(ha, "center")
                cell.border    = _thin_border()
                if fmt != "General":
                    cell.number_format = fmt

            ws.row_dimensions[row].height = 16
            row += 1
    else:
        ws.merge_cells(f"B{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D, "[ alarm_by_ne.csv not found ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]), align=_align("left"))
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1   # spacer

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-SECTION C: Alarm Category Breakdown
    # ─────────────────────────────────────────────────────────────────────────
    _section_bar(ws, row, D, D + NC - 1,
                 "  C.  ALARM CATEGORY BREAKDOWN  "
                 "(Radio · Transport · Power · Hardware · Software)")
    row += 1

    _col_headers(ws, row, D,
                 ["#", "Alarm Category", "Count", "% of Total",
                  "Avg Duration (min)", "Severity Mix"],
                 height=22)
    row += 1

    cat_df = data.get("alarm_category")
    if cat_df is not None and not cat_df.empty:
        for i, (_, r) in enumerate(cat_df.iterrows()):
            cat = r.get("alarm_category", "")
            bg  = _CAT_FILL.get(cat, (_C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]))

            vals = [
                i+1,
                cat,
                r.get("count", ""),
                r.get("pct_of_total", ""),
                r.get("avg_duration_min", ""),
                str(r.get("severity_mix", "")),
            ]
            fmts   = ["General", "General", "#,##0", "0.00", "0.0", "General"]
            aligns = ["center", "left", "center", "center", "center", "left"]

            for j, (v, fmt, ha) in enumerate(zip(vals, fmts, aligns)):
                bold  = (j == 1)
                txt_c = _C["NAVY_DARK"] if j == 1 else _C["TEXT_DARK"]
                cell  = ws.cell(row=row, column=D+j, value=v)
                cell.fill      = _fill(bg)
                cell.font      = _font(bold=bold, size=9, color=txt_c)
                cell.alignment = _align(ha, "center", wrap=(j == 5))
                cell.border    = _thin_border()
                if fmt != "General":
                    cell.number_format = fmt

            ws.row_dimensions[row].height = 16
            row += 1
    else:
        ws.merge_cells(f"B{row}:{get_column_letter(D+NC-1)}{row}")
        _write(ws, row, D, "[ alarm_by_category.csv not found ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]), align=_align("left"))
        row += 1

    ws.freeze_panes = "B5"


# ═══════════════════════════════════════════════════════════════════════════════
# 7.  SHEET 5 — KPI BREACH SITE RANKING
# ═══════════════════════════════════════════════════════════════════════════════

def _build_kpi_site_sheet(ws, data: dict):
    """
    Populate the KPI Breach Site Ranking worksheet.

    Source: breach_summary_per_site.csv
    Top 20 sites by any-breach rate; per-metric breach columns rendered
    dynamically (whichever metric columns exist in the CSV).
    """
    ws.title = "KPI Site Ranking"
    ws.sheet_properties.tabColor = _TAB["site"]
    ws.sheet_view.showGridLines  = True

    D = 2   # data start column

    # ── Rows 1–2: title banner ────────────────────────────────────────────────
    _title_banner(ws, 1, 2, 1, 18,
                  "HFCL EMS — KPI Breach Analysis: Site-Level Ranking",
                  font_size=15)

    ws.row_dimensions[3].height = 14
    ws.merge_cells("B3:P3")
    cap = ws["B3"]
    cap.value = (
        "Top 20 network elements ranked by any-KPI breach rate (fraction of 15-min windows "
        "where at least one KPI crossed its SLA threshold).  "
        "Per-metric columns show individual KPI breach rates.  "
        "Red ≥30% · Amber 15–30% · Alternating blue-grey otherwise."
    )
    cap.font      = _font(size=9, italic=True, color=_C["TEXT_MUTED"])
    cap.alignment = _align("left", "center", wrap=True)
    ws.row_dimensions[4].height = 8   # spacer

    site_df = data.get("breach_site")

    if site_df is None or site_df.empty:
        ws.merge_cells("B5:P5")
        _write(ws, 5, D,
               "[ breach_summary_per_site.csv not found — run Phase 4 KPI trending analysis ]",
               font=_font(italic=True, color=_C["TEXT_MUTED"]),
               align=_align("left"))
        return

    # Sort by any_breach_pct
    if "any_breach_pct" in site_df.columns:
        site_df = site_df.sort_values("any_breach_pct", ascending=False)

    # Detect per-metric breach columns dynamically
    metric_cols   = sorted([c for c in site_df.columns
                             if c.endswith("_breach_pct") and c != "any_breach_pct"])
    metric_labels = [_kpi_label(c.replace("_breach_pct", "")) for c in metric_cols]

    # Set column widths
    fixed_widths = {1: 3, 2: 6, 3: 22, 4: 10, 5: 10, 6: 14, 7: 16}
    for c, w in fixed_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for i in range(len(metric_cols)):
        ws.column_dimensions[get_column_letter(8 + i)].width = 14
    # Right margin
    ws.column_dimensions[get_column_letter(8 + len(metric_cols))].width = 3

    # Column headers
    static_hdrs = [
        "#", "Network Element", "Tech", "Region",
        "Observations", "Any-Breach\nRate (%)",
    ]
    metric_hdrs = [f"{ml}\nBreach (%)" for ml in metric_labels]
    _col_headers(ws, 5, D, static_hdrs + metric_hdrs, height=38)

    row = 6
    for i, (_, r) in enumerate(site_df.head(20).iterrows()):
        any_r = r.get("any_breach_pct", None)
        if isinstance(any_r, (int, float)) and not pd.isna(any_r):
            bg = (_C["CRIT_BG"] if any_r >= 30 else
                  _C["WARN_BG"] if any_r >= 15 else
                  (_C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]))
        else:
            bg = _C["ROW_EVEN"] if i % 2 == 0 else _C["ROW_ODD"]

        static_vals = [
            i+1,
            r.get("network_element_id", ""),
            r.get("technology", ""),
            r.get("region", ""),
            r.get("n_observations", ""),
            any_r,
        ]
        static_fmts = ["General", "General", "General", "General", "#,##0", "0.00"]

        for j, (v, fmt) in enumerate(zip(static_vals, static_fmts)):
            col_idx = D + j
            bold    = (j == 1)
            txt_c   = _C["NAVY_DARK"] if j == 1 else _C["TEXT_DARK"]
            halign  = "left" if j == 1 else "center"
            cell    = ws.cell(row=row, column=col_idx, value=v)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=bold, size=9, color=txt_c)
            cell.alignment = _align(halign, "center")
            cell.border    = _thin_border()
            if fmt != "General":
                cell.number_format = fmt

        # Per-metric columns
        for k, mc in enumerate(metric_cols):
            v       = r.get(mc, None)
            col_idx = D + len(static_vals) + k
            # Individual metric cell highlight
            if isinstance(v, (int, float)) and not pd.isna(v):
                cell_bg = (_C["CRIT_BG"] if v >= 25 else
                           _C["WARN_BG"] if v >= 12 else bg)
            else:
                cell_bg = bg

            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.fill         = _fill(cell_bg)
            cell.font         = _font(size=9)
            cell.alignment    = _align("center", "center")
            cell.border       = _thin_border()
            cell.number_format = "0.00"

        ws.row_dimensions[row].height = 16
        row += 1

    ws.freeze_panes = "B6"


# ═══════════════════════════════════════════════════════════════════════════════
# 8.  MAIN PUBLIC API
# ═══════════════════════════════════════════════════════════════════════════════

def generate_report(
    output_dir: str | Path = "output",
    phase4_reports_subpath: str = "reports/phase4",
) -> Path:
    """
    Generate ``Weekly_Network_Health_Summary.xlsx`` inside ``output_dir/reports/``.

    Parameters
    ----------
    output_dir
        Root output folder (default: ``output/`` at project root).
        The Phase 4 CSVs are expected at
        ``<output_dir>/<phase4_reports_subpath>/``.
    phase4_reports_subpath
        Subdirectory under ``output_dir`` where Phase 4 CSVs live.
        Default: ``reports/phase4``.

    Returns
    -------
    ``pathlib.Path`` pointing to the written ``.xlsx`` file.
    """
    output_dir   = Path(output_dir)
    reports_dir  = output_dir / phase4_reports_subpath
    excel_outdir = output_dir / "reports"
    excel_outdir.mkdir(parents=True, exist_ok=True)
    out_path = excel_outdir / "Weekly_Network_Health_Summary.xlsx"

    log.info("=" * 65)
    log.info("PHASE 5 — Corporate Report Generator")
    log.info("=" * 65)
    log.info(f"  Phase 4 CSVs  : {reports_dir}")
    log.info(f"  Output file   : {out_path}")
    log.info("")

    # 1. Load Phase 4 analytical CSVs ─────────────────────────────────────────
    log.info("Loading Phase 4 analytical CSVs…")
    data = _load_csvs(reports_dir)

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d  %H:%M UTC")

    # 2. Initialise workbook ───────────────────────────────────────────────────
    log.info("\nBuilding workbook sheets…")
    wb = Workbook()

    # openpyxl creates a default sheet — use it as Sheet 1
    ws_exec = wb.active
    log.info("  Sheet 1 → Executive Summary")
    _build_executive_summary(ws_exec, data, generated_at)

    log.info("  Sheet 2 → Root-Cause Candidates")
    ws_rc = wb.create_sheet()
    _build_root_cause_sheet(ws_rc, data)

    log.info("  Sheet 3 → Degradation Windows")
    ws_deg = wb.create_sheet()
    _build_degradation_sheet(ws_deg, data)

    log.info("  Sheet 4 → Alarm Frequency")
    ws_alarm = wb.create_sheet()
    _build_alarm_frequency_sheet(ws_alarm, data)

    log.info("  Sheet 5 → KPI Site Ranking")
    ws_site = wb.create_sheet()
    _build_kpi_site_sheet(ws_site, data)

    # 3. Workbook-level metadata ───────────────────────────────────────────────
    wb.properties.title   = "HFCL EMS — Weekly Network Health Summary"
    wb.properties.creator = "HFCL EMS Phase 5 Reporting Engine"
    wb.properties.subject = "4G/5G Network KPI & Alarm Analytics"
    wb.properties.description = (
        "Auto-generated weekly corporate report. "
        "Data sourced from Phase 4 CSV exports in output/reports/phase4/."
    )

    # 4. Save ──────────────────────────────────────────────────────────────────
    wb.save(out_path)
    log.info(f"\n✓  Report saved → {out_path}")
    log.info("=" * 65)

    return out_path


# ═══════════════════════════════════════════════════════════════════════════════
# 9.  CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def _cli():
    parser = argparse.ArgumentParser(
        prog="python -m phase5_reporting.report_generator",
        description="HFCL EMS Phase 5 — Weekly Network Health Report Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples\n"
            "--------\n"
            "  python -m phase5_reporting.report_generator\n"
            "  python -m phase5_reporting.report_generator --output-dir output\n"
            "  python -m phase5_reporting.report_generator --output-dir /data/ems/reports\n"
            "\n"
            "The report is written to:\n"
            "  <output-dir>/reports/Weekly_Network_Health_Summary.xlsx\n"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="output",
        metavar="PATH",
        help="Base output directory that contains reports/phase4/ "
             "(default: output/)",
    )
    parser.add_argument(
        "--phase4-subpath",
        type=str,
        default="reports/phase4",
        metavar="SUBPATH",
        help="Subdirectory under --output-dir where Phase 4 CSVs live "
             "(default: reports/phase4)",
    )
    args = parser.parse_args()

    try:
        out = generate_report(
            output_dir=args.output_dir,
            phase4_reports_subpath=args.phase4_subpath,
        )
        print(f"\n  ✓  Report written to: {out}\n")
        sys.exit(0)
    except Exception as exc:
        import traceback
        log.error(f"Report generation failed: {exc}")
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    _cli()